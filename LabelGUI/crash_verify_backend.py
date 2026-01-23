import os
import cv2
import time
import csv
from datetime import datetime, timedelta
import json
import video_utils

# Single active session is enough for your demo and avoids complexity.
_STATE = {
    "sid": None,
    "video_path": None,
    "person": None,
    "scenario": None,
    "youtube_link": None,
    "pred_crash_events": None,
    "pred_crashes_per_min": None,
    "verified_crash_events": 0,
    "verified_times_sec": [],
    "duration_sec": 0.0,
    "done": False,
    "saved": False,
    "last_error": None,
}

def _repo_root():
    # LabelGUI/crash_verify_backend.py -> repo root is one level up
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def _analysis_results_dir():
    return os.path.join(_repo_root(), "analysis", "results")

def start_crash_verify_session(sid, video_path, person, scenario, youtube_link, predicted_crash_events="", predicted_crashes_per_min=""):
    global _STATE
    _STATE = {
        "sid": sid,
        "video_path": video_path,
        "person": person,
        "scenario": scenario,
        "youtube_link": youtube_link,
        "pred_crash_events": predicted_crash_events,
        "pred_crashes_per_min": predicted_crashes_per_min,
        "verified_crash_events": 0,
        "verified_times_sec": [],
        "duration_sec": 0.0,
        "done": False,
        "saved": False,
        "last_error": None,
    }
    # reset shared time tracking
    video_utils.set_current_time_sec(0.0)
    video_utils.set_video_duration(0.0)

def mark_crash_now(sid):
    if _STATE.get("sid") != sid:
        return
    t = float(video_utils.get_current_time_sec() or 0.0)
    _STATE["verified_crash_events"] = int(_STATE.get("verified_crash_events", 0)) + 1
    _STATE["verified_times_sec"].append(t)

def get_crash_verify_status(sid):
    if _STATE.get("sid") != sid:
        return {"ok": False, "error": "Invalid session."}
    duration = float(_STATE.get("duration_sec") or 0.0)
    verified = int(_STATE.get("verified_crash_events") or 0)
    verified_per_min = (verified / (duration / 60.0)) if duration > 0 else 0.0

    return {
        "ok": True,
        "done": bool(_STATE.get("done")),
        "saved": bool(_STATE.get("saved")),
        "person": _STATE.get("person"),
        "scenario": _STATE.get("scenario"),
        "youtube_link": _STATE.get("youtube_link"),
        "pred_crash_events": _STATE.get("pred_crash_events"),
        "pred_crashes_per_min": _STATE.get("pred_crashes_per_min"),
        "verified_crash_events": verified,
        "verified_crashes_per_min": round(verified_per_min, 3),
        "duration_sec": round(duration, 2),
        "last_error": _STATE.get("last_error"),
    }

def generate_crash_verify_stream(sid):
    if _STATE.get("sid") != sid:
        yield b""
        return

    video_path = _STATE.get("video_path")
    if not video_path or not os.path.exists(video_path):
        _STATE["last_error"] = f"Video not found: {video_path}"
        _STATE["done"] = True
        yield b""
        return

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        _STATE["last_error"] = f"OpenCV could not open video: {video_path}"
        _STATE["done"] = True
        yield b""
        return

    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    total_frames = cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0.0
    duration = (total_frames / fps) if total_frames > 0 else 0.0
    _STATE["duration_sec"] = float(duration)

    video_utils.set_video_duration(duration)
    video_utils.set_current_time_sec(0.0)

    def draw_overlay(frame, current_time_sec):
        elapsed_str = str(timedelta(seconds=int(current_time_sec)))
        total_str = str(timedelta(seconds=int(duration)))
        verified = int(_STATE.get("verified_crash_events", 0))
        pred = _STATE.get("pred_crash_events", "")

        cv2.putText(frame, f"{elapsed_str} / {total_str}", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
        cv2.putText(frame, f"Verified crashes: {verified}", (10, 70),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 200, 255), 2)
        if str(pred).strip() != "":
            cv2.putText(frame, f"Predicted crashes: {pred}", (10, 110),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 200, 0), 2)
        return frame

    for mjpeg_frame in video_utils.read_video_frames(cap, fps, draw_overlay):
        if not mjpeg_frame:
            time.sleep(0.05)
            continue
        yield mjpeg_frame

    _STATE["done"] = True

def _append_verified_csv(row_dict):
    results_dir = _analysis_results_dir()
    os.makedirs(results_dir, exist_ok=True)
    path = os.path.join(results_dir, "verified.csv")
    exists = os.path.exists(path)

    header = [
        "person", "scenario", "youtube_link",
        "verified_crash_events", "verified_crashes_per_min",
        "verified_times_sec", "timestamp", "notes"
    ]

    with open(path, "a", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header)
        if not exists:
            w.writeheader()
        w.writerow({k: row_dict.get(k, "") for k in header})

def _append_to_excel_sheet(excel_path, row_dict):
    """
    Append to a sheet named 'CrashVerification' inside the uploaded Excel.
    Creates the sheet if missing.
    """
    if not excel_path or not os.path.exists(excel_path):
        return

    try:
        import openpyxl
    except Exception:
        return

    wb = openpyxl.load_workbook(excel_path)
    sheet_name = "CrashVerification"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    header = [
        "person", "scenario", "youtube_link",
        "pred_crash_events", "pred_crashes_per_min",
        "verified_crash_events", "verified_crashes_per_min",
        "verified_times_sec", "timestamp", "notes"
    ]

    # write header if sheet is empty
    if ws.max_row == 1 and all(ws.cell(1, c).value in (None, "") for c in range(1, len(header)+1)):
        ws.append(header)

    ws.append([row_dict.get(k, "") for k in header])
    wb.save(excel_path)

def finish_crash_verify_session(sid, excel_path=None, notes=""):
    """
    Persist results to analysis/results/verified.csv and also append to uploaded Excel (optional).
    """
    if _STATE.get("sid") != sid:
        return False, "Invalid session."

    if _STATE.get("saved"):
        return True, "Already saved."

    duration = float(_STATE.get("duration_sec") or 0.0)
    verified = int(_STATE.get("verified_crash_events") or 0)
    verified_per_min = (verified / (duration / 60.0)) if duration > 0 else 0.0

    row = {
        "person": _STATE.get("person"),
        "scenario": _STATE.get("scenario"),
        "youtube_link": _STATE.get("youtube_link"),
        "pred_crash_events": _STATE.get("pred_crash_events"),
        "pred_crashes_per_min": _STATE.get("pred_crashes_per_min"),
        "verified_crash_events": verified,
        "verified_crashes_per_min": round(verified_per_min, 6),
        "verified_times_sec": json.dumps([round(float(t), 3) for t in _STATE.get("verified_times_sec", [])]),
        "timestamp": datetime.utcnow().isoformat(),
        "notes": notes,
    }

    try:
        _append_verified_csv(row)
    except Exception as e:
        return False, f"Failed to save verified.csv: {e}"

    try:
        _append_to_excel_sheet(excel_path, row)
    except Exception:
        # Excel write is "best effort" — csv is the real source of truth
        pass

    _STATE["saved"] = True
    return True, "Saved verification results."
